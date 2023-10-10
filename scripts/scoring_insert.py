import requests 
import json
from openpyxl import load_workbook
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)
file = load_workbook('{rutaPadre}descargas/{asunto}/{carnetDash}/scoring_{carnetDash}.xlsx',data_only=True)
#file = load_workbook('D:/AD/BUROS_05/excels/plantilla_scoring.xlsx',data_only=True)
sheet = file.active
lista = []
for row in range(5,sheet.max_row+1):
    cell = sheet.cell(row=row, column=3)
    lista.append(str(cell.value))
for row in range(5,sheet.max_row+1):
    cell = sheet.cell(row=row, column=4)
    lista.append(str(cell.value))
print(lista)

persona = GetVar('datoDash')
persona = persona.strip("()")

# Divide la cadena en elementos separados por comas y crea una lista
persona = persona.split(', ')

# Evalúa cada elemento de la lista para manejar tipos diferentes
for i in range(len(persona)):
    elemento = persona[i]
    if elemento.isdigit():  # Utiliza isdigit() para verificar enteros
        persona[i] = int(elemento)
    elif elemento.replace(".", "", 1).isdigit():  # Verifica si es un número decimal
        persona[i] = float(elemento)
    elif elemento.startswith("'") and elemento.endswith("'"):  # Si es una cadena, quita las comillas
        persona[i] = elemento.strip("'")
print(persona)
##################################
        ##TIPO##
##################################
data = "{datoDash}"
split_data = data.split(", ")
tipo= split_data[4]
tipo = tipo.strip("'")
print(" TIPO :",tipo)

ticket = GetVar('ticket_db')
robot = GetVar('numeroRobot')
robot = 'Robot buros '+str(robot)

#url = 'http://10.10.92.18/backend/api/RBT_SCORING/Scoring/Insertar'  ###api/Robot/Scoring/Insertar' #'http://localhost/APIECOFUTURO/api/Robot/Scoring/Insertar'
#['A', '1', 'None', '0', '0', '1', '5', '0', '>10000', 'No', '0', 'None', '>15', 'No', 'None', 'None', 'None', '150', '150', 'None', '62.5', '12.5', '62.5', '62.5', '62.5', '62.5', '62.5', '62.5', 'None', '0', '100', '850', 'None', 'Bajo']
url = 'http://localhost/ApiBuros/apiBuros/Scoring/Insert'
datos ={
  "ticket": str(ticket),
  "nroDocumento": str(persona[2]),
  "complemento": str(persona[10]),
  "extension": str(persona[9]),
  "usuario": robot,
  "tipo": tipo,
  "cicResCalif": str(lista[0]),
  "cicPunCalif": str(lista[17]),
  "cicResNumEifPef": str(lista[1]),
  "cicPunNumEifPef": str(lista[18]),
  "credResNumPef": str(lista[3]),
  "credPunNumPef": str(lista[20]),
  "credResNumEifDirPef": str(lista[4]),
  "credPunNumEifDirPef": str(lista[21]),
  "credResHistCalifDir60mes": str(lista[5]),
  "credPunHistCalifDir60mes": str(lista[22]),
  "credResHistCalifIndir60mes": str(lista[6]),
  "credPunHistCalifIndir60mes": str(lista[23]),
  "credResDeudasCasaComs": str(lista[7]),
  "credPunDeudasCasaComs": str(lista[24]),
  "credResMontoOperaciones": str(lista[8]),
  "credPunMontoOperaciones": str(lista[25]),
  "credResEif90d": str(lista[9]),
  "credPunEif90d": str(lista[26]),
  "credResNroEifConsultadas": str(lista[10]),
  "credPunNroEifConsultadas": str(lista[27]),
  "hojaResRiesgosAtraso": str(lista[12]),
  "hojaPunRiesgosAtraso": str(lista[29]),
  "hojaResRiesgosCondonaciones": str(lista[13]),
  "hojaPunRiesgosCondonaciones": str(lista[30]),
  "totalPuntaje": str(lista[31]),
  "nivelRiesgo": str(lista[33])
}

print (datos)
payload = json.dumps(datos)
headers = {'Content-Type': 'application/json'}
response = requests.request("POST", url, headers=headers, data=payload, verify=False)
print(response.text)
##################################
        ##DatosJSON##
##################################
ruta = r'{rutaPadre}scoring'
nombre_archivo = 'scoring.txt'

# Crear la carpeta "scoring" si no existe
if not os.path.exists(ruta):
    os.makedirs(ruta)
ruta_completa = os.path.join(ruta, nombre_archivo)
with open(ruta_completa, 'w') as file:
    file.write("\n\nDato Lista Excel :\n")
    json.dump(lista, file, indent=4)
    file.write("Datos de la variable persona:\n")
    json.dump(persona, file, indent=4)
    file.write("\n\nDatos del JSON:\n")
    json.dump(datos, file, indent=4)
    file.write("\n\nDato TIPO:\n")
    json.dump("{tipo}", file, indent=4)
    file.write("\n\nDato personas_db:\n")
    json.dump("{personas_db}", file, indent=4)
    file.write("\n\nDato datoDash:\n")
    json.dump("{datoDash}", file, indent=4)
    file.write("\n\nDato TIPO :\n")
    json.dump(tipo, file, indent=4)

print(f"Archivo '{nombre_archivo}' creado en '{ruta_completa}'.")
