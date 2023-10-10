import requests 
import json
from openpyxl import load_workbook
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)
file = load_workbook('{rutaPadre}descargas/{asunto}/{carnetDash}/scoring_{carnetDash}.xlsx',data_only=True)
sheet = file.active
lista = []
for row in range(5,sheet.max_row+1):
    cell = sheet.cell(row=row, column=3)
    lista.append(str(cell.value))
for row in range(5,sheet.max_row+1):
    cell = sheet.cell(row=row, column=4)
    lista.append(str(cell.value))
print(lista)

#persona = [21497, 10696, '5764793', 'ALEXANDER', 'TITULAR', '', 'CARNET DE IDENTIDAD', 'MORANTE', 'COTA', '', '', '', '', 0] #GetVar('datoDash')
ticket = GetVar('ticket_db')
robot = GetVar('numeroRobot')
robot = 'Robot buros '+str(robot)

url = 'http://10.10.92.18/backend/api/Robot/Scoring/Insertar' #'http://localhost/APIECOFUTURO/api/Robot/Scoring/Insertar'
#['A', '1', 'None', '0', '0', '1', '5', '0', '>10000', 'No', '0', 'None', '>15', 'No', 'None', 'None', 'None', '150', '150', 'None', '62.5', '12.5', '62.5', '62.5', '62.5', '62.5', '62.5', '62.5', 'None', '0', '100', '850', 'None', 'Bajo']

datos ={
  "NRO_DOCUMENTO": persona[2],
  "COMPLEMENTO": persona[10],
  "EXTENSION": persona[9],
  "TICKET": ticket,
  "USUARIO": robot,
  "CIC": [
    {
      "CRITERIO": "Calificacion",
      "RESULTADO": str(lista[0]),
      "PUNTAJE": str(lista[17])
    },
    {
      "CRITERIO": "Numero de EIF Directas incluyendo a PEF",
      "RESULTADO": str(lista[1]),
      "PUNTAJE": str(lista[18])
    }
  ],
  "Infocred": [
    {
      "CRITERIO": "Numero de operaciones indirectas incluyendo PEF",
      "RESULTADO": str(lista[3]),
      "PUNTAJE": str(lista[20])
    },
    {
      "CRITERIO": "Numero de EIF Directas incluyendo a PEF",
      "RESULTADO": str(lista[4]),
      "PUNTAJE": str(lista[21])
    },
    {
      "CRITERIO": "Historial - Calificacion directa ultimos 60 meses",
      "RESULTADO": str(lista[5]),
      "PUNTAJE": str(lista[22])
    },
    {
      "CRITERIO": "Historial - Calificacion indirecta ultimos 60 meses",
      "RESULTADO": str(lista[6]),
      "PUNTAJE": str(lista[23])
    },
    {
      "CRITERIO": "Deudas en casa comerciales",
      "RESULTADO": str(lista[7]),
      "PUNTAJE": str(lista[24])
    },
    {
      "CRITERIO": "Monto de la operaciones",
      "RESULTADO": str(lista[8]),
      "PUNTAJE": str(lista[25])
    },
    {
      "CRITERIO": "Consulta en otras EIF en los ultimos 90 dias",
      "RESULTADO": str(lista[9]),
      "PUNTAJE": str(lista[26])
    },
    {
      "CRITERIO": "Nro de EIF consultadas",
      "RESULTADO": str(lista[10]),
      "PUNTAJE": str(lista[27])
    }
  ],
  "Hoja_Riesgos": [
    {
      "CRITERIO": "Maximo de dias de atraso",
      "RESULTADO": str(lista[12]),
      "PUNTAJE": str(lista[29])
    },
    {
      "CRITERIO": "Condonaciones",
      "RESULTADO": str(lista[13]),
      "PUNTAJE": str(lista[30])
    }
  ],
  "Puntaje": [
    {
      "CRITERIO": "Total Puntaje",
      "RESULTADO": "0",
      "PUNTAJE": str(lista[31])
    },{
      "CRITERIO": "Nivel de riesgo",
      "RESULTADO": "0",
      "PUNTAJE": str(lista[33])
    }
  ]
}
print (datos)
payload = json.dumps(datos)
headers = {'Content-Type': 'application/json'}
response = requests.request("POST", url, headers=headers, data=payload, verify=False)
print(response.text)
